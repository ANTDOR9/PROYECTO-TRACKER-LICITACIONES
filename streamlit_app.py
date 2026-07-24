"""
TRACKER DE LICITACIONES - Brighter Peru   (version NUBE / Streamlit Cloud)
Oportunidades VIGENTES con el Estado, en vivo desde la API publica de SEACE.
NUEVO: etiquetas de busqueda EDITABLES desde la propia pagina (sin tocar codigo).
"""
import os, re, unicodedata
from io import BytesIO
import pandas as pd
import requests
import streamlit as st
import yaml

API = ("https://prod4.seace.gob.pe:8086/api/oportunidades/"
       "codObjeto/codDepartamento/sintesisProceso/codTipoProceso/0/0/0/0")

DEF_CLAVES = ["pantalla interactiva","pizarra digital","monitor interactivo",
              "pantalla led","senalizacion digital","video wall","kiosco",
              "totem","equipamiento audiovisual","proyector"]
DEF_EXCLUIR = ["protector de pantalla","luminaria","presa","drone","optotipo"]

def separar(txt):
    """Separa etiquetas por coma o por salto de linea (ambos modos)."""
    return [x.strip() for x in re.split(r"[,\n]", txt) if x.strip()]

def norm(t):
    if not t: return ""
    t = unicodedata.normalize("NFKD", str(t))
    return "".join(c for c in t if not unicodedata.combining(c)).lower()

def cargar_config():
    ruta = os.path.join(os.path.dirname(__file__), "config.yaml")
    claves, excluir = DEF_CLAVES, DEF_EXCLUIR
    if os.path.exists(ruta):
        c = yaml.safe_load(open(ruta, encoding="utf-8")) or {}
        pk = c.get("palabras_clave", {})
        if isinstance(pk, dict):
            claves = [t for terms in pk.values() for t in terms]
        elif isinstance(pk, list):
            claves = pk
        excluir = c.get("palabras_excluir", DEF_EXCLUIR)
    return claves, excluir

@st.cache_data(ttl=3600)
def traer_todo():
    """Trae TODAS las oportunidades vigentes (sin filtrar) para filtrar en vivo."""
    import urllib3; urllib3.disable_warnings()
    r = requests.get(API, timeout=120, verify=False)
    r.raise_for_status()
    data = r.json()
    filas = []
    for o in data:
        texto = norm(f"{o.get('detObjeto','')} {o.get('detItem','')} {o.get('sintesisProceso','')} {o.get('nomenclatura','')}")
        valor = pd.to_numeric(o.get("valorReferencial") or o.get("valorReferencialItem"), errors="coerce")
        filas.append({
            "_texto": texto,
            "Nomenclatura": o.get("nomenclatura",""),
            "Entidad": o.get("detEntidad",""),
            "Objeto": o.get("detObjeto",""),
            "Descripcion": o.get("detItem","") or o.get("sintesisProceso",""),
            "Valor referencial": valor,
            "Tipo": o.get("detTipoProceso",""),
            "Fin inscripcion": o.get("fechaFin",""),
            "Presentacion propuestas": o.get("fechaPresentacionPropuestas",""),
        })
    return pd.DataFrame(filas), len(data)

def etiqueta_de(texto, etiquetas, excluir):
    if any(x in texto for x in excluir): return None
    for et in etiquetas:
        toks = [t for t in norm(et).split() if len(t) >= 4]
        if toks and all(t in texto for t in toks):
            return et
    return None

def exportar_excel(df):
    """Genera un Excel ordenado y facil de leer (bytes)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    orden = ["Etiqueta","Nomenclatura","Entidad","Objeto","Descripcion",
             "Valor referencial","Tipo","Fin inscripcion","Presentacion propuestas"]
    cols = [c for c in orden if c in df.columns]
    d = df[cols].copy()
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        d.to_excel(writer, index=False, sheet_name="Oportunidades", startrow=1)
        wb = writer.book
        ws = writer.sheets["Oportunidades"]
        # Titulo
        ws.cell(row=1, column=1, value="Tracker de Licitaciones - Brighter Peru")
        ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="1F3B4D")
        # Encabezados (fila 2)
        fill = PatternFill("solid", fgColor="1F3B4D")
        thin = Side(style="thin", color="D9D9D9")
        borde = Border(left=thin, right=thin, top=thin, bottom=thin)
        for j, col in enumerate(cols, start=1):
            c = ws.cell(row=2, column=j)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = borde
        # Anchos por columna
        anchos = {"Etiqueta":22,"Nomenclatura":26,"Entidad":34,"Objeto":16,
                  "Descripcion":50,"Valor referencial":18,"Tipo":22,
                  "Fin inscripcion":18,"Presentacion propuestas":22}
        for j, col in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(j)].width = anchos.get(col, 18)
        # Formato de moneda + bordes + zebra en las filas de datos
        n = len(d)
        for i in range(n):
            fila = 3 + i
            for j, col in enumerate(cols, start=1):
                c = ws.cell(row=fila, column=j)
                c.border = borde
                c.alignment = Alignment(vertical="top", wrap_text=(col in ("Descripcion","Entidad")))
                if col == "Valor referencial":
                    c.number_format = '"S/" #,##0.00'
                if i % 2 == 1:
                    c.fill = PatternFill("solid", fgColor="F2F5F7")
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{2+n}"
    return buffer.getvalue()

st.set_page_config(page_title="Tracker de Licitaciones - Brighter", page_icon="📡", layout="wide")
st.title("📡 Tracker de Licitaciones — Brighter Perú")
st.caption("Oportunidades VIGENTES con el Estado (a las que postular ahora). Fuente: SEACE / OECE (en vivo).")

try:
    df, total = traer_todo()
except Exception as e:
    st.error(f"No se pudo consultar la API de SEACE.\n\n{e}")
    st.stop()

claves_cfg, excluir_cfg = cargar_config()

# --- PANEL EDITABLE DE ETIQUETAS ---
st.sidebar.header("🏷️ Etiquetas de búsqueda")
st.sidebar.caption("Separa por coma o por línea. Edita para investigar otros productos.")
txt_etiquetas = st.sidebar.text_area("Etiquetas de interés",
    value="\n".join(claves_cfg), height=180, key="etq")
txt_excluir = st.sidebar.text_area("Excluir (una por línea)",
    value="\n".join(excluir_cfg), height=90, key="exc")

etiquetas = separar(txt_etiquetas)
excluir = [norm(e) for e in separar(txt_excluir)]

# clasificar segun etiquetas editadas
df = df.copy()
df["Etiqueta"] = df["_texto"].apply(lambda t: etiqueta_de(t, etiquetas, excluir))
rel = df[df["Etiqueta"].notna()].drop(columns=["_texto"])

st.sidebar.divider()
st.sidebar.header("Filtros")
vmin = st.sidebar.number_input("Valor referencial mínimo (S/)", 0, step=1000, value=0)
buscar = st.sidebar.text_input("Buscar en descripción / entidad")

f = rel.copy()
if vmin: f = f[f["Valor referencial"].fillna(0) >= vmin]
if buscar:
    t = buscar.lower()
    f = f[f["Descripcion"].str.lower().str.contains(t, na=False) |
          f["Entidad"].str.lower().str.contains(t, na=False)]

c1, c2, c3 = st.columns(3)
c1.metric("Oportunidades encontradas", len(f))
c2.metric("Etiquetas activas", len(etiquetas))
c3.metric("Procesos revisados", f"{total:,}")

if f.empty:
    st.warning("Ninguna oportunidad vigente coincide con las etiquetas actuales. "
               "Prueba agregando o cambiando términos en el panel de la izquierda.")
else:
    st.divider()
    st.dataframe(
        f[["Etiqueta","Nomenclatura","Entidad","Objeto","Descripcion",
           "Valor referencial","Tipo","Fin inscripcion","Presentacion propuestas"]]
          .sort_values("Fin inscripcion"),
        use_container_width=True, hide_index=True)

    with st.expander("Resumen por etiqueta"):
        st.dataframe(f.groupby("Etiqueta").size().reset_index(name="Oportunidades"),
                     use_container_width=True, hide_index=True)

    st.download_button("⬇️ Descargar Excel",
        exportar_excel(f),
        file_name="tracker_licitaciones.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
